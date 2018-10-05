#grabs clients from spreadsheet, creates client objects, gets credits from month sheet
#adds credits to clients and prints

import openpyxl
from flask import Flask,render_template
from tdd2 import open_booking_wb, find_tab, cell_has_value
import templates



class Client:

    def __init__(self, first, last):
        self.firstN = first
        self.lastN = last
        self.credits = 0

    @property
    def fullname(self):
        return f'{self.firstN} {self.lastN}'

    def add_credits(self, num):
        self.credits += num


# def get_rates(book):
#     rates = {}
#     ws = find_tab(book, "Rates")
#     rows = ws.iter_rows(min_col=2, min_row=2, max_col=3)
#     for row in rows:
#         rates[row[0].value] = row[1].value
#     return rates


def create_dict(ws, keycol, valuecol):
    dictionary = {}
    keys = ws[keycol]
    values = ws[valuecol]

    length = len(keys)
    for i in range(1, length):
        dictionary[keys[i].value] = values[i].value
    return dictionary


def create_clients(wb):
    client_list = {}
    ws = find_tab(wb, 'Clients')

    for cell in ws['A'][1:]:
        sp = cell.value.split()
        client = Client(sp[0], sp[1])
        client_list[client.firstN] = client

    return client_list


def fill_credits(wb, date, client_list):
    faculties = create_dict(wb['Facilities'], 'A', 'B')
    rates = create_dict(wb['Rates'], 'B', 'C')

    datesheet = wb[date]

    for cell in datesheet[1]:
        if cell_has_value(cell):
            column = cell.column
            roomname = cell.value
            rate = rates[faculties[roomname]]

            for name in datesheet[column][1:]:
                if cell_has_value(name):
                    client_list[name.value].add_credits(rate)


def create_report(wb, date, client_list):
    report = []
    fill_credits(wb, date, client_list)

    for client in client_list:
        report.append(f'Name: {client_list[client].firstN} | Credits: {client_list[client].credits} \n')

    return report


def create_report_for_html(wb, date, client_list):
    report = []
    fill_credits(wb, date, client_list)

    for client in client_list:
        report.append((client_list[client].fullname, client_list[client].credits))

    return report


workb = open_booking_wb('cSpace_Bookingv1.xlsx')

app = Flask(__name__)


@app.route('/')
def index():
    return 'Index'

# def report_html(date):
#     client_list = create_clients(workb)
#     rep = create_report_for_html(workb, date, client_list)
#
#     report_string = """
#     <!DOCTYPE html>
#     <html>
#     <h2> Clients and Credits for """ + date + """</h2>
#             <table border = '1'>
#                 <tr>
#                     <td>
#                         Name
#                     </td>
#                     <td>
#                         Credits
#                     </td>
#                 </tr>
#                     """
#     for i in rep:
#         report_string += "<tr>"
#         report_string += "<td>"
#         report_string += i[0]
#         report_string += "</td>"
#         report_string += "<td>"
#         report_string += str(i[1])
#         report_string += "</td>"
#         report_string += "<tr>"
#     report_string += "</table></h2></html>"
#     return report_string


@app.route('/<date>')
def report_html(date):
    client_list = create_clients(workb)
    rep = create_report_for_html(workb, date, client_list)

    return render_template('clientrates.html', rep=rep)
