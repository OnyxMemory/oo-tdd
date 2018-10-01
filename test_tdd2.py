import unittest
import tdd2
import datetime
import openpyxl
from openpyxl import Workbook, worksheet


class Test_tdd2(unittest.TestCase):
    def test_add_five(self):
        self.assertEqual(8, tdd2.add_five(3))

    def test_my_max(self):
        self.assertEqual(5, tdd2.my_max([1, 2, 3, 4, 5]))

    def test_my_min(self):
        self.assertEqual(1, tdd2.my_min([1, 2, 3, 4, 5]))

    def test_has_string(self):
        self.assertEqual(["Mary had"], tdd2.has_string(["Mary had",
                                                        "a little lamb",
                                                        "little lamb",
                                                        "Whose fleece", ],
                                                       "Mary"))

    def test_to_date(self):
        dt = tdd2.to_date("2010-08-02")
        self.assertIsInstance(dt, datetime.date)
        self.assertEqual(2010, dt.year)
        self.assertEqual(8, dt.month)
        self.assertEqual(2, dt.day)

    def test_date_diff(self):
        self.assertEqual(1, tdd2.date_diff("2018-09-02", "2018-09-01"))
        self.assertEqual(2080, tdd2.date_diff("2018-09-01", "2012-12-21"))

    def test_contains(self):
        self.assertTrue(tdd2.contains(['a', 'b', 'd'], "a"))
        self.assertFalse(tdd2.contains(['c', 'b', 'd', 'f'], "a"))

    def test_add_contents(self):
        self.assertEqual(6, tdd2.add_contents([1, 2, 3]))

    def test_lookup(self):
        self.assertEqual('one mine', tdd2.lookup({1: 'one', 2: 'two', 3: 'three'}, 1))

    def test_get_clients(self):
        wb = self.create_test_wb()
        self.assertEqual(["John", "Sally", "Harry", "Bob"],
                         tdd2.get_clients(wb))

    def test_find_tab(self):
        wb = self.create_test_wb()
        entered_date = '2012-07'

        ws = tdd2.find_tab(wb, entered_date)
        self.assertIsInstance(ws, worksheet.Worksheet)
        self.assertEqual(entered_date[0:7], ws.title)

        self.assertIsNone(tdd2.find_tab(wb, '2018-07'))

    def test_string_in_array(self):
        self.assertTrue(tdd2.string_in_array('a', ['a', 'b', 'c']))
        self.assertFalse(tdd2.string_in_array('d', ['e', 'f', 'g']))

    def test_verify_clients_month(self):
        wb = self.create_test_wb()
        # tdd2.verify_clients_month(wb,"2012-07"
        self.assertEqual(["Badguy"], tdd2.verify_clients_month(wb, "2012-07"))

    @staticmethod
    def create_test_wb():
        wb = Workbook()
        wb.create_sheet("Mysheet")
        wb.create_sheet("Mysheet2")
        wb.create_sheet("2012-07")
        wb.create_sheet("2012-08")
        wb.create_sheet("2012-09")
        wb.create_sheet("Clients")

        ws = wb["2012-07"]
        ws.cell(row=1, column=3).value = 'Room a'
        ws.cell(row=1, column=4).value = 'Room b'
        ws.cell(row=1, column=5).value = 'Room c'

        ws.cell(row=2, column=3).value = 'Harry'
        ws.cell(row=2, column=4).value = 'Harry'
        ws.cell(row=2, column=5).value = None

        ws.cell(row=5, column=3).value = 'Harry'
        ws.cell(row=5, column=4).value = 'Badguy'

        ws = wb["Clients"]
        ws.cell(row=1, column=1).value = "Name"
        ws.cell(row=2, column=1).value = "John"
        ws.cell(row=3, column=1).value = "Sally"
        ws.cell(row=4, column=1).value = "Harry"
        ws.cell(row=5, column=1).value = "Bob"

        # wb.save('test.xlsx')

        return wb
