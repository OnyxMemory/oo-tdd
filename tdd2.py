from datetime import datetime
import sys
import openpyxl
from openpyxl import load_workbook


# from openpyxl import Workbook

def add_five(num):
    return num + 5


def my_max(array):
    if not array:
        return None
    mx = array[0]
    length = len(array)
    for i in range(0, length):
        if mx < array[i]:
            mx = array[i]
    return mx


def my_min(array):
    if not array:
        return None
    mn = array[0]
    length = len(array)
    for i in range(0, length):
        if mn > array[i]:
            mn = array[i]
    return mn


def has_string(array, string):
    stringArray = []
    for entry in array:
        if string in entry:
            stringArray.append(entry)
    return stringArray


def to_date(date):
    return datetime.strptime(date, "%Y-%m-%d").date()


def date_diff(date1, date2):
    firstdate = to_date(date1)
    seconddate = to_date(date2)
    # print((firstdate - seconddate).days)

    return abs((firstdate - seconddate).days)


def contains(array, value):
    for i in array:
        if value == i:
            return True
    return False


def add_contents(array):
    total = 0
    for i in array:
        total += i
    return total


def lookup(diction, key):
    if key not in diction:
        return " mine"
    else:
        return diction[key] + " mine"


def find_tab(wb, tab):
    for sheet in wb:
        if sheet.title == tab:
            return sheet
    return None


def find_tab_date(wb, date_s):
    for sheet in wb:
        if sheet.title == date_s[0:7]:
            return sheet
    return None


def open_booking_wb(name):
    return load_workbook(name)


def get_clients(wb):
    # wb = open_booking_wb(name)
    ws = find_tab(wb, "Clients")
    names = []
    for cell in ws["A"][1:]:
        names.append(cell.value)
    # print (names)
    return names


def get_room_data(ws):
    return ws.iter_rows(min_col=3, min_row=2)


def cell_has_value(cell):
    if not cell.value:
        return False
    if not cell.value.strip():
        return False
    return True


def string_in_array(check, array):
    for item in array:
        if check in item:
            return True
    return False


def verify_clients_month(wb, month):
    non_clients = []
    ws = find_tab(wb, month)
    rooms = get_room_data(ws)
    for row in rooms:
        for cell in row:
            if cell_has_value(cell):
                if not string_in_array(cell.value, get_clients(wb)):
                    if not string_in_array(cell.value, non_clients):
                        non_clients.append(cell.value)
    return non_clients
