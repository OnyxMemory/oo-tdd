import unittest
import clientrates
import tdd2
from openpyxl import Workbook, load_workbook


class TestClientrates(unittest.TestCase):

    def test_init(self):
        client1 = clientrates.Client('Test', 'User')

        self.assertIsInstance(client1, clientrates.Client)

    def test_create_dict(self):
        wb = self.create_test_wb()
        ws = wb["Rates"]
        rates = clientrates.create_dict(ws, 'B', 'C')
        self.assertEqual(rates['Desk'], 1)
        self.assertEqual(rates['Gallery'], 5)

    def test_create_clients(self):
        wb = self.create_test_wb()
        clientlist = clientrates.create_clients(wb)
        self.assertEqual('Bob Not', clientlist['Bob'].fullname)

    def test_fill_credits(self):
        wb = self.create_test_wb()
        client_list = clientrates.create_clients(wb)
        clientrates.fill_credits(wb, '2012-07', client_list)

        self.assertEqual(2, client_list['Harry'].credits)
        self.assertEqual(4, client_list['Sally'].credits)

    def test_print_report(self):
        wb = self.create_test_wb()
        client_list = clientrates.create_clients(wb)
        report = clientrates.create_report(wb, '2012-07', client_list)

        # for i in report:
        #     print(i)

    def test_excel_report(self):
        wb = load_workbook('cSpace_Bookingv1.xlsx')
        client_list = clientrates.create_clients(wb)
        report = clientrates.create_report(wb, '2018-10', client_list)

        for i in report:
            print(i)



    @staticmethod
    def create_test_wb():
        wb = Workbook()
        wb.create_sheet("Mysheet")
        wb.create_sheet("Mysheet2")
        wb.create_sheet("2012-07")
        wb.create_sheet("2012-08")
        wb.create_sheet("2012-09")
        wb.create_sheet("Clients")
        wb.create_sheet("Rates")
        wb.create_sheet("Facilities")

        ws = wb["2012-07"]
        ws.cell(row=1, column=3).value = 'Room a'
        ws.cell(row=1, column=4).value = 'Room b'
        ws.cell(row=1, column=5).value = 'Room c'

        ws.cell(row=2, column=3).value = 'Harry'
        ws.cell(row=2, column=4).value = 'Bob'
        ws.cell(row=2, column=5).value = None

        ws.cell(row=5, column=3).value = 'Harry'
        ws.cell(row=5, column=4).value = 'Sally'
        ws.cell(row=5, column=5).value = 'John'

        ws.cell(row=6, column=3).value = 'Sally'

        ws = wb["Clients"]
        ws.cell(row=1, column=1).value = "Name"
        ws.cell(row=2, column=1).value = "John Bon"
        ws.cell(row=3, column=1).value = "Sally Doe"
        ws.cell(row=4, column=1).value = "Harry Snow"
        ws.cell(row=5, column=1).value = "Bob Not"

        ws = wb["Rates"]
        ws.cell(row=2, column=2).value = "Desk"
        ws.cell(row=2, column=3).value = 1
        ws.cell(row=3, column=2).value = "Meeting"
        ws.cell(row=3, column=3).value = 3
        ws.cell(row=4, column=2).value = "Gallery"
        ws.cell(row=4, column=3).value = 5

        ws = wb["Facilities"]
        ws.cell(row=2, column=1).value = 'Room a'
        ws.cell(row=2, column=2).value = 'Desk'
        ws.cell(row=3, column=1).value = 'Room b'
        ws.cell(row=3, column=2).value = 'Meeting'
        ws.cell(row=4, column=1).value = 'Room c'
        ws.cell(row=4, column=2).value = 'Gallery'

        # wb.save('test.xlsx')
        return wb

